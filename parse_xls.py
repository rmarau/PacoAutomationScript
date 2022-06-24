from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from pathlib import Path

from models import Sumario


EXCEL_SHEET_PLANNING = "Planeamento"
EXCEL_SHEET_PLANNING_UC_CODE = "B1"
EXCEL_SHEET_PLANNING_TP_CODE = "B2"
EXCEL_SHEET_PLANNING_DRY_RUN = "B3"
EXCEL_SHEET_PLANNING_LINE_STARTS_AT = 6
EXCEL_SHEET_PLANNING_COL_DATE = "A"
EXCEL_SHEET_PLANNING_COL_HORAS = "B"
EXCEL_SHEET_PLANNING_COL_AULA = "C"
EXCEL_SHEET_PLANNING_COL_SALA = "D"
EXCEL_SHEET_PLANNING_COL_STATUS = "E"
EXCEL_SHEET_PLANNING_COL_SUMARIO = "F"
EXCEL_SHEET_PLANNING_COL_BIBLIOGRAFIA = "G"
EXCEL_SHEET_PUBLISHED_LABEL = "PUBLISHED"

EXCEL_SHEET_ATTENDANCE = "PresençaAulas"
EXCEL_SHEET_ATTENDANCE_LINHA_SEM_CABECALHOS = 4
EXCEL_SHEET_ATTENDANCE_LINHA_AULAS = 2
EXCEL_SHEET_ATTENDANCE_COLUNA_MECS = 'A'

class Excel:

    def __init__(self, nome, pagina=EXCEL_SHEET_PLANNING, pagina_faltas=EXCEL_SHEET_ATTENDANCE):
        self._nome = nome
        self._sheet = pagina
        self._sheet_faltas = pagina_faltas

        #check if closed
        xls = Path(nome)
        xls_tmp = xls.parent/('~$'+xls.name)

        while xls_tmp.is_file():
            print(xls_tmp, 'is opened!...   ' + 'Close it and press Enter!    (You cannot update an excel file while open)')
            input()

    def get_uc_code(self):
        wb = load_workbook(self._nome, read_only=True, data_only=True)
        ws = wb.active if self._sheet is None else wb[self._sheet]
        uc_code = str(ws[EXCEL_SHEET_PLANNING_UC_CODE].value)
        wb.close()
        return uc_code

    def get_tp_code(self):
        wb = load_workbook(self._nome, read_only=True, data_only=True)
        ws = wb.active if self._sheet is None else wb[self._sheet]
        tp_code = str(ws[EXCEL_SHEET_PLANNING_TP_CODE].value)
        wb.close()
        return tp_code

    def get_dry_run_bool(self):
        wb = load_workbook(self._nome, read_only=True, data_only=True)
        ws = wb.active if self._sheet is None else wb[self._sheet]
        dry_run_bool = ws[EXCEL_SHEET_PLANNING_DRY_RUN].value
        wb.close()
        return dry_run_bool

    def get_presencas_mec(self, aula):
        wb = load_workbook(self._nome, data_only=True)
        ws = wb.active if self._sheet_faltas is None else wb[self._sheet_faltas]

        #Search Column
        column = None
        for cell in ws[f'A{EXCEL_SHEET_ATTENDANCE_LINHA_AULAS}:ZZ{EXCEL_SHEET_ATTENDANCE_LINHA_AULAS}'][0]:
            if cell.value == aula:
                column = get_column_letter(cell.column)
                break

        if column is not None:
            #for cell in ws[column]: print(cell)
            start_offset = EXCEL_SHEET_ATTENDANCE_LINHA_SEM_CABECALHOS - 1
            check_column = ws[column][start_offset:]
            yield from (
                ws[f'{EXCEL_SHEET_ATTENDANCE_COLUNA_MECS}{cell.row}'].value
                for cell in check_column if cell.value is not None and not (cell.value==0 or cell.value=="0" or cell.value==False)
            )

        wb.close()

    def get_sumarios_bulk(self):
        wb = load_workbook(self._nome, read_only=True, data_only=True)
        ws = wb.active if self._sheet is None else wb[self._sheet]

        i = EXCEL_SHEET_PLANNING_LINE_STARTS_AT
        while True:
            date_datetime = ws[f'{EXCEL_SHEET_PLANNING_COL_DATE}{i}'].value
            if date_datetime is None: date_datetime = ws[f'{EXCEL_SHEET_PLANNING_COL_DATE}{i-1}'].value

            hora = ws[f'{EXCEL_SHEET_PLANNING_COL_HORAS}{i}'].value
            aula = ws[f'{EXCEL_SHEET_PLANNING_COL_AULA}{i}'].value
            sala = ws[f'{EXCEL_SHEET_PLANNING_COL_SALA}{i}'].value
            status = ws[f'{EXCEL_SHEET_PLANNING_COL_STATUS}{i}'].value
            sumario = ws[f'{EXCEL_SHEET_PLANNING_COL_SUMARIO}{i}'].value
            bibliografia = ws[f'{EXCEL_SHEET_PLANNING_COL_BIBLIOGRAFIA}{i}'].value

            if hora is None: break

            yield Sumario(date_datetime, hora, aula, status, sumario, bibliografia, sala,
                          list( str(mec) for mec in self.get_presencas_mec(aula) )
                          )
            i += 1

        wb.close()

    def get_sumarios_hours_filtered(self):
        """Filter out Sumários with zero hours or None"""
        return ( s for s in self.get_sumarios_bulk() if s.hora is not None and int(s.hora)>0 )

    def get_sumarios_marked_published(self):
        return ( s for s in self.get_sumarios_hours_filtered() if s.status is not None and s.status == EXCEL_SHEET_PUBLISHED_LABEL)

    def get_sumarios_to_publish(self):
        """Filter out Sumários already published (marked as)"""
        return ( s for s in self.get_sumarios_hours_filtered() if s.status is None or s.status != EXCEL_SHEET_PUBLISHED_LABEL)

    def get_sumarios_to_publish_with_attendance_filled(self):
        """ Filter sumarios with at least one student attending.
            Hopefully ther's at least one student checked """
        return ( s for s in self.get_sumarios_to_publish() if s.presencas_mec )


    def update_status_published(self, in_aulas):
        #First compile the lines to update.
        #I have to do this search in data_only! and with so, I cannot save
        #because I'll loose formulas.
        lines_to_update = []

        wb = load_workbook(self._nome, read_only=True, data_only=True)
        ws = wb.active if self._sheet is None else wb[self._sheet]

        i = EXCEL_SHEET_PLANNING_LINE_STARTS_AT
        while True:
            hora = ws[f'{EXCEL_SHEET_PLANNING_COL_HORAS}{i}'].value
            aula = ws[f'{EXCEL_SHEET_PLANNING_COL_AULA}{i}'].value
            if aula in in_aulas: lines_to_update.append(i)
            if hora is None: break
            i += 1
        wb.close()

        #Now We can go and update the status
        wb = load_workbook(self._nome, read_only=False, data_only=False)
        ws = wb.active if self._sheet is None else wb[self._sheet]

        for line in lines_to_update: ws[f'{EXCEL_SHEET_PLANNING_COL_STATUS}{line}'].value = EXCEL_SHEET_PUBLISHED_LABEL

        #wb.save(filename="t.xlsx")
        wb.save(filename=self._nome)
        wb.close()

